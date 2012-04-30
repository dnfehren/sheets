#!/usr/bin/env python

import os, collections, itertools, re

import pprint
pp = pprint.PrettyPrinter(indent=4)

#Intro

'''
Wrapper class for reading common spreadsheet types. 
CSV, XLS and XLSX are supported currently

Dan Fehrenbach
dnfehrenbach@gmail.com

dependencies are collections (for named tuples), csv, xlrd and openpyxl

why use tuple for rows?
http://stackoverflow.com/questions/626759/whats-the-difference-between-list-and-tuples-in-python

row are heterogeneous, right? but a row, when its with other rows is homogeneous
so... 
rows are tuples (sometimes named tuples)
worksheets are named tuples containing a name and data (data is a list of tuples)
workbooks are lists containing worksheets
'''

#Work
#Functions

def clean_header_cell(potential_header_cell):
    '''
    takes a string and returns a string with problematic parts removed
    or replaced, to ensure compatability with python keys, databases
    '''
    rx_spec_chars = re.compile("['(',')','$','-']")
    rx_space = re.compile("\s+")
    #rx_first_digit = re.compile("^\d") #used with named tuples

    level_0 = rx_spec_chars.sub('', str(potential_header_cell))
    level_1 = rx_space.sub('',level_0)
   
    return level_1  


def make_header(potential_header_row):
    '''
    check a row from a spreadsheet and create a list suitable for use as
    '''
    clean_header = []

    for p_head_num, p_head_cell in enumerate(potential_header_row):
       
        if p_head_cell == '':
            clean_header.append('col' + str(p_head_num))
        else:
            clean_header.append(clean_header_cell(p_head_cell))

    return clean_header


def xlsx_row_values(openpyxl_row):
    '''
    take a row from the openpyxl iterator and return just the values
     this (likely) depends on the row being generated from the iter_rows()
     method that comes with using an openpyxl reader with use_iterators=True
    '''
    values_only_row = []

    for cell in openpyxl_row:
        values_only_row.append(cell.internal_value)

    return values_only_row




class SheetReader(object):

    def __init__(self, file_location, header=[True], header_row=[1]):

        try:
            file_handle = open(file_location, 'rb')
        except IOError:
            print "can't open file"


        file_basename = os.path.basename(file_location)
        file_name, file_ext = os.path.splitext(file_basename)

        
        worksheet_template = {'sheet_name':'','sheet_headers':'','sheet_data':''}
        
        '''
        CSV Handling
        '''
        if file_ext == '.csv':

            import csv

            csv_reader = csv.reader(file_handle)

            '''
            TODO
            #what about really big sheets?
            file_size = os.path.getsize(file_location)

            if file_size > some_limit:
                #then what? can't access csv by row/col indexs
            '''

            working_sheet = [] #holds the row tuples
            header_list = []

            for row_num, row in enumerate(csv_reader):
                if header is True:
                    if row_num == header_row - 1: #enumerate starts at 0
                        header_list = make_header(row)
                    elif row_num < header_row - 1:
                        continue
                    else:
                        working_sheet.append(tuple(row))
                else:
                    '''
                    with no header expected a basic tuple of the row
                    is appended to the list of rows
                    ''' 
                    working_sheet.append(tuple(row))

            '''
            the workbook and worksheet construction is used here,
            even though csv's won't have more than one worksheet, to maintain
            consistency in functions that will access the sheet data later
            '''
            workbook = [] #for non-csv, a list of worksheets, here it's only 1
            
            worksheet = worksheet_template.copy()

            worksheet['sheet_name'] = file_name
            worksheet['sheet_headers'] = header_list
            worksheet['sheet_data'] = working_sheet

            workbook.append(worksheet)
            
            self.sheets = workbook
        
        
        elif file_ext == '.xls':
            
            import xlrd

            try:
                xl_reader = xlrd.open_workbook(file_location)
            except xlrd.biffh.XLRDError:
                print "excel problem on file load"
                

            xl_sheet_names = xl_reader.sheet_names()

            workbook = []

            for xl_sheet_index, xl_sheet_name in enumerate(xl_sheet_names):

                try:
                    sheet_header_flag = header[xl_sheet_index]
                except IndexError:
                    sheet_header_flag = True

                try:
                    sheet_header_row = header_row[xl_sheet_index]
                except IndexError:
                    sheet_header_row = 1

                working_sheet = []
                sheet_header = []
                
                xl_sheet = xl_reader.sheet_by_name(xl_sheet_name)

                '''
                to get xlrd row values you need an index number,
                AFAIK there isnt a way to iterate by row w/o an index
                '''
                for xl_row_num in range(xl_sheet.nrows):

                    if sheet_header_flag is True:
                        if xl_row_num < sheet_header_row - 1:
                            continue
                        elif xl_row_num == sheet_header_row - 1:          
                            p_head_row = xl_sheet.row_values(xl_row_num)
                            sheet_header = make_header(p_head_row)
                        else:
                            xl_row = xl_sheet.row_values(xl_row_num)
                            working_sheet.append(tuple(xl_row))
                    else:
                        xl_row = xl_sheet.row_values(xl_row_num)
                        working_sheet.append(tuple(xl_row))

                worksheet = worksheet_template.copy()
                worksheet['sheet_name'] = xl_sheet_name
                worksheet['sheet_headers'] = sheet_header
                worksheet['sheet_data'] = working_sheet

                workbook.append(worksheet)    

            self.sheets = workbook

        elif file_ext == '.xlsx':

            import openpyxl

            xlsx_reader = openpyxl.load_workbook(filename = file_location, 
                                                use_iterators=True)

            xlsx_sheet_names = xlsx_reader.get_sheet_names()

            workbook = []

            for xlsx_sheet_index, xlsx_sheet_name in enumerate(xlsx_sheet_names):

                try:
                    sheet_header_flag = header[xlsx_sheet_index]
                except IndexError:
                    sheet_header_flag = True

                try:
                    sheet_header_row = header_row[xlsx_sheet_index]
                except IndexError:
                    sheet_header_row = 1

                working_sheet = []
                sheet_header = []

                xlsx_sheet = xlsx_reader.get_sheet_by_name(name = xlsx_sheet_name)

                '''
                openpyxl doesnt have a row values function, cells need to have
                their values extracted individually, using sheet.iter_rows() 
                should make things as fast as possible
                REF: http://packages.python.org/openpyxl/optimized.html
                '''
                for xlsx_row_num, xlsx_row in enumerate(xlsx_sheet.iter_rows()):

                    if sheet_header_flag is True:

                        if xlsx_row_num == sheet_header_row - 1:
                            p_head_row = xlsx_row_values(xlsx_row)
                            sheet_header = make_header(p_head_row)
                        else:
                            xlsx_row = xlsx_row_values(xlsx_row)
                            working_sheet.append(tuple(xlsx_row))
                    else:
                        xlsx_row = xlsx_row_values(xlsx_row)
                        working_sheet.append(tuple(xlsx_row))

                worksheet = worksheet_template.copy()
                worksheet['sheet_name'] = xlsx_sheet_name
                worksheet['sheet_headers'] = sheet_header
                worksheet['sheet_data'] = working_sheet


                workbook.append(worksheet)

            self.sheets = workbook
        
        else:
            print "unsupported file type. Only .csv, .xls and .xlsx"


    def book_names(self):
        '''
        returns the names of all worksheets in a workbook
        '''
        b_names = [sht['sheet_name'] for sht in self.sheets]
        return b_names

    
    def sheet_rows(self, *args):
        '''
        returns an iterator over the rows of the  specified sheets in the workbook

        can be used in the following ways...
         with no sheets specified, you get all rows from all sheets
         with one sheet specified, you get the rows from that sheet
         with multiple sheets spec'd, you get all the row from those sheets
        '''
        if len(args) == 0:
            for sheet in self.sheets:
                for row in sheet['sheet_data']:
                    yield row
        elif len(args) == 1:
            for sheet in self.sheets:
                if sheet['sheet_name'] == args[0]:
                    for row in sheet['sheet_data']:
                        yield row
        else:
            wb_names = self.book_names()
            sheet_matches = [s_try for s_try in args if s_try in wb_names]
            for sheet_name in sheet_matches:
                for sheet in self.sheets:
                    if sheet['sheet_name'] == sheet_name:
                        for row in sheet['sheet_data']:
                            yield row
     
    
    def sheet_cols(self, *args):
        '''
        returns an iterator over the columns in specifed sheets in the workbook

        can be used in the following ways...
         with no sheets specified you get all cols from all sheets
         with one sheet specified you get the cols from that sheet
         with more than one sheet sped'd you get the cols from those sheets
        ''' 
        if len(args) == 0:
            for sheet in self.sheets:
                for col in itertools.izip_longest(*sheet['sheet_data']):
                    yield col
        elif len(args) == 1:
            for sheet in self.sheets:
                if sheet['sheet_name'] == args[0]:
                    for col in itertools.izip_longest(*sheet['sheet_data']):
                        yield col
        else:
            wb_names = self.book_names()
            sheet_matches = [s_try for s_try in args if s_try in wb_names]
            for sheet_name in sheet_matches:
                for sheet in self.sheets:
                    if sheet['sheet_name'] == sheet_name:
                        for col in itertools.izip_longest(*sheet['sheet_data']):
                            yield col
        
    '''
    #http://code.activestate.com/recipes/192401-quickly-remove-or-order-columns-in-a-list-of-lists/
    #http://stackoverflow.com/questions/1983902/remove-row-or-column-from-2d-list-if-all-values-in-that-row-or-column-are-none
    def del_col(self, sheet_name, col_num_list):
        for sheet in self.sheets:
            if sheet.sheet_name == sheet_name:
                pass
                #go through each row and delete something at specific index

                #delete a whole column
    '''


    def convert_to_databook(self):
        
        try:
            import tablib
        except ImportError:
            print "problems importing tablib, install it please"

        tblb_databook = tablib.Databook()

        for sheet in self.sheets:
            pass
            #tblb_dataset = tablib.Dataset(self.sheet_data)
            #tblb_databook


    def convert_to_dataframe(self, sheet_name):

        #TODO
        # should this asssume that this has been imported in another
        # script before the function runs, put a try catch
        # around the DataFrame creation line?
        try:
            import pandas as pandas
        except ImportError:
            print "problems importing pandas"

        col_sets = []

        for sheet in self.sheets:
            if sheet['sheet_name'] == sheet_name:
                for col_num, col in enumerate(self.sheet_cols(sheet_name)):
                    col_data = (sheet['sheet_headers'][col_num],col)
                    col_sets.append(col_data)

        df = pandas.DataFrame.from_items(col_sets)

        return df
         


